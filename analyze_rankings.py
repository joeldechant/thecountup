"""
Analyze TheCountUp Rankings data and produce:
  - CountUp Analysis.xlsx       (multi-sheet summary)
  - analytics/analytics.html    (standalone styled page)
  - analytics/analytics_chart.png (embedded bar chart)

Usage: python analyze_rankings.py  (Excel must be open)
"""
import xlwings as xw
import os, statistics, html
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

ANALYTICS_DIR = os.path.join(os.path.dirname(__file__), "analytics")


def clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    return s if s != "." else ""


# ---------------------------------------------------------------------------
# Data extraction (mirrors build_website.py logic)
# ---------------------------------------------------------------------------

def extract_candy(ws):
    data = ws.range((1, 1), (ws.used_range.last_cell.row, ws.used_range.last_cell.column)).value
    items = []
    for i, row in enumerate(data):
        if i + 1 < 3:
            continue
        country = clean(row[1]) if len(row) > 1 else ""
        brand = clean(row[2]) if len(row) > 2 else ""
        variety = clean(row[3]) if len(row) > 3 else ""
        price = row[5] if len(row) > 5 and isinstance(row[5], (int, float)) else None
        if not country or not brand:
            continue
        if (variety.lower() == "total") or (brand.lower() == "total"):
            continue
        items.append({"country": country, "brand": brand, "variety": variety, "price": price})
    return items


def extract_chocolate(ws):
    data = ws.range((1, 1), (ws.used_range.last_cell.row, ws.used_range.last_cell.column)).value
    items = []
    for i, row in enumerate(data):
        if i + 1 < 4:
            continue
        brand = clean(row[1])
        name = clean(row[2])
        choc_type = clean(row[3])
        if not brand or not name or brand == "Top Chocolate":
            continue
        ct = choc_type.strip().capitalize() if choc_type else ""
        if ct.startswith("Dark"):
            ct = "Dark"
        elif ct.lower().startswith("milk") or ct == "MIlk":
            ct = "Milk"
        cacao = row[4] if isinstance(row[4], (int, float)) else None
        sugar = row[5] if isinstance(row[5], (int, float)) else None
        items.append({"brand": brand, "name": name, "type": ct, "cacao": cacao, "sugar": sugar})
    return items


def extract_pop(ws):
    data = ws.range((1, 1), (ws.used_range.last_cell.row, ws.used_range.last_cell.column)).value
    items = []
    for i, row in enumerate(data):
        if i + 1 < 3:
            continue
        raw_b = row[1]
        if raw_b is not None and str(raw_b).strip() == ".":
            break
        brand = clean(row[1])
        flavor = clean(row[2])
        if not brand or not flavor:
            continue
        if brand in ("Avg", "Adds", "Offset", "Offset Owed"):
            break
        sugar = row[3] if isinstance(row[3], (int, float)) else None
        items.append({"brand": brand, "flavor": flavor, "sugar": sugar})
    return items


def extract_sauces(ws):
    data = ws.range((1, 1), (ws.used_range.last_cell.row, ws.used_range.last_cell.column)).value
    section_names = {"Asian Sauces", "Other Main Sauces", "Finishing Sauces"}
    items = []
    for i, row in enumerate(data):
        if i + 1 < 3:
            continue
        col_b = clean(row[1])
        col_c = clean(row[2])
        if not col_b and col_c in section_names:
            continue
        origin = col_b
        name = col_c
        brand = clean(row[3])
        if not origin or not name or not brand:
            continue
        rice = clean(row[4])
        chicken = clean(row[5])
        items.append({"origin": origin, "name": name, "brand": brand, "rice": rice, "chicken": chicken})
    return items


def extract_dining(ws):
    data = ws.range((1, 1), (ws.used_range.last_cell.row, ws.used_range.last_cell.column)).value
    items = []
    for i, row in enumerate(data):
        if i + 1 < 3:
            continue
        restaurant = clean(row[1])
        item = clean(row[2])
        rating = clean(row[3])
        hack_price = row[7] if isinstance(row[7], (int, float)) else None
        if not restaurant or not item:
            continue
        items.append({"restaurant": restaurant, "item": item, "rating": rating, "hack_price": hack_price})
    return items


def extract_games(ws):
    data = ws.range((1, 1), (ws.used_range.last_cell.row, ws.used_range.last_cell.column)).value
    items = []
    seen = set()
    for i, row in enumerate(data):
        if i + 1 < 3:
            continue
        rank_val = row[1]
        game = clean(row[2])
        if not game or not isinstance(rank_val, (int, float)):
            continue
        if game in seen:
            continue
        seen.add(game)
        avg = row[6] if isinstance(row[6], (int, float)) else None
        year = int(row[8]) if len(row) > 8 and isinstance(row[8], (int, float)) else None
        items.append({"game": game, "year": year, "score": avg})
    return items


# ---------------------------------------------------------------------------
# Analysis helpers
# ---------------------------------------------------------------------------

GRADE_ORDER = ["A++", "A+", "A", "B+", "B", "C+", "C", "D+", "D", "F"]


def grade_distribution(grades):
    counts = Counter(g.strip() for g in grades if g.strip() and g.strip() != "\u2014")
    return {g: counts.get(g, 0) for g in GRADE_ORDER if counts.get(g, 0) > 0}


def price_stats(prices):
    valid = [p for p in prices if p is not None]
    if not valid:
        return {}
    return {
        "count": len(valid),
        "min": min(valid),
        "max": max(valid),
        "avg": statistics.mean(valid),
        "median": statistics.median(valid),
    }


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws_xl):
    for cell in ws_xl[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws_xl):
    for col in ws_xl.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws_xl.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def write_excel(all_data):
    wb = Workbook()

    # --- Overview ---
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Category", "Items"])
    for name, items in all_data.items():
        ws.append([name, len(items)])
    ws.append(["TOTAL", sum(len(v) for v in all_data.values())])
    style_header(ws)
    auto_width(ws)

    # --- Geography ---
    ws = wb.create_sheet("Geography")
    geo = Counter()
    for item in all_data["Candy"]:
        geo[item["country"]] += 1
    for item in all_data["Sauces"]:
        geo[item["origin"]] += 1
    ws.append(["Country / Origin", "Candy", "Sauces", "Total"])
    candy_geo = Counter(i["country"] for i in all_data["Candy"])
    sauce_geo = Counter(i["origin"] for i in all_data["Sauces"])
    all_regions = sorted(set(list(candy_geo.keys()) + list(sauce_geo.keys())),
                         key=lambda x: -(candy_geo[x] + sauce_geo[x]))
    for region in all_regions:
        ws.append([region, candy_geo[region], sauce_geo[region], candy_geo[region] + sauce_geo[region]])
    style_header(ws)
    auto_width(ws)

    # --- Price Analysis ---
    ws = wb.create_sheet("Price Analysis")
    ws.append(["--- Candy $/4oz by Country ---", "", "", "", ""])
    ws.append(["Country", "Count", "Min", "Max", "Avg", "Median"])
    candy_by_country = defaultdict(list)
    for item in all_data["Candy"]:
        if item["price"] is not None:
            candy_by_country[item["country"]].append(item["price"])
    for country in sorted(candy_by_country, key=lambda c: statistics.mean(candy_by_country[c])):
        s = price_stats(candy_by_country[country])
        ws.append([country, s["count"], f"${s['min']:.2f}", f"${s['max']:.2f}",
                    f"${s['avg']:.2f}", f"${s['median']:.2f}"])
    ws.append([])
    ws.append(["--- Fast Food Hack Prices ---", "", "", ""])
    ws.append(["Restaurant", "Item", "Rating", "Hack Price"])
    for item in sorted(all_data["Fast Food Hack"], key=lambda x: x["hack_price"] or 999):
        p = f"${item['hack_price']:.2f}" if item["hack_price"] else ""
        ws.append([item["restaurant"], item["item"], item["rating"], p])
    for row in ws.iter_rows(min_row=2, max_row=2, max_col=6):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
    auto_width(ws)

    # --- Grades ---
    ws = wb.create_sheet("Grades")
    ws.append(["--- Sauces Grade Distribution ---", "", ""])
    ws.append(["Grade", "Rice", "Chicken"])
    rice_dist = grade_distribution([i["rice"] for i in all_data["Sauces"]])
    chicken_dist = grade_distribution([i["chicken"] for i in all_data["Sauces"]])
    all_grades = sorted(set(list(rice_dist.keys()) + list(chicken_dist.keys())),
                        key=lambda g: GRADE_ORDER.index(g) if g in GRADE_ORDER else 99)
    for g in all_grades:
        ws.append([g, rice_dist.get(g, 0), chicken_dist.get(g, 0)])
    ws.append([])
    ws.append(["--- Fast Food Hack Ratings ---", ""])
    ws.append(["Rating", "Count"])
    rating_dist = grade_distribution([i["rating"] for i in all_data["Fast Food Hack"]])
    for g, c in rating_dist.items():
        ws.append([g, c])
    for row in ws.iter_rows(min_row=2, max_row=2, max_col=3):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
    auto_width(ws)

    # --- Pop Sugar ---
    ws = wb.create_sheet("Pop Sugar")
    ws.append(["--- Sugar by Brand ---", "", "", ""])
    ws.append(["Brand", "Count", "Avg Sugar (g)", "Min", "Max"])
    pop_by_brand = defaultdict(list)
    for item in all_data["Pop"]:
        if item["sugar"] is not None:
            pop_by_brand[item["brand"]].append(item["sugar"])
    for brand in sorted(pop_by_brand, key=lambda b: -statistics.mean(pop_by_brand[b])):
        vals = pop_by_brand[brand]
        ws.append([brand, len(vals), f"{statistics.mean(vals):.1f}",
                    f"{min(vals):.0f}", f"{max(vals):.0f}"])
    ws.append([])
    ws.append(["--- Top 5 Highest Sugar ---", "", ""])
    ws.append(["Brand", "Flavor", "Sugar (g)"])
    pop_with_sugar = [i for i in all_data["Pop"] if i["sugar"] is not None]
    for item in sorted(pop_with_sugar, key=lambda x: -x["sugar"])[:5]:
        ws.append([item["brand"], item["flavor"], f"{item['sugar']:.0f}"])
    ws.append([])
    ws.append(["--- Top 5 Lowest Sugar ---", "", ""])
    ws.append(["Brand", "Flavor", "Sugar (g)"])
    for item in sorted(pop_with_sugar, key=lambda x: x["sugar"])[:5]:
        ws.append([item["brand"], item["flavor"], f"{item['sugar']:.0f}"])
    for row in ws.iter_rows(min_row=2, max_row=2, max_col=5):
        for cell in row:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
    auto_width(ws)

    # --- Games by Decade ---
    ws = wb.create_sheet("Games by Decade")
    ws.append(["Decade", "Count", "Avg Score"])
    decade_data = defaultdict(list)
    for item in all_data["Games"]:
        if item["year"] and item["score"]:
            decade = (item["year"] // 10) * 10
            decade_data[decade].append(item["score"])
    for decade in sorted(decade_data):
        scores = decade_data[decade]
        ws.append([f"{decade}s", len(scores), f"{statistics.mean(scores):.2f}"])
    style_header(ws)
    auto_width(ws)

    path = os.path.join(os.path.dirname(__file__), "CountUp Analysis.xlsx")
    wb.save(path)
    print(f"Written: {path}")
    return path


# ---------------------------------------------------------------------------
# Chart
# ---------------------------------------------------------------------------

def make_chart(all_data):
    geo = Counter()
    for item in all_data["Candy"]:
        geo[item["country"]] += 1
    for item in all_data["Sauces"]:
        geo[item["origin"]] += 1

    top = geo.most_common(15)
    labels = [t[0] for t in reversed(top)]
    values = [t[1] for t in reversed(top)]

    fig, ax = plt.subplots(figsize=(10, 7))
    fig.patch.set_facecolor("#000000")
    ax.set_facecolor("#000000")
    bars = ax.barh(labels, values, color="#ffffff", edgecolor="#333333", height=0.6)
    ax.set_xlabel("Items Ranked", color="#ffffff", fontsize=12, fontfamily="monospace")
    ax.set_title("Top 15 Countries / Origins (Candy + Sauces)", color="#ffffff",
                 fontsize=14, fontfamily="serif", fontweight="bold", pad=15)
    ax.tick_params(colors="#ffffff", labelsize=10)
    for spine in ax.spines.values():
        spine.set_color("#333333")
    ax.xaxis.label.set_fontfamily("monospace")
    for label in ax.get_yticklabels():
        label.set_fontfamily("monospace")
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height() / 2,
                str(val), va="center", ha="left", color="#ffffff",
                fontsize=10, fontfamily="monospace")

    os.makedirs(ANALYTICS_DIR, exist_ok=True)
    path = os.path.join(ANALYTICS_DIR, "analytics_chart.png")
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor="#000000")
    plt.close(fig)
    print(f"Written: {path}")
    return path


# ---------------------------------------------------------------------------
# HTML analytics page
# ---------------------------------------------------------------------------

def esc(text):
    return html.escape(str(text))


def build_html(all_data):
    candy_prices = [i["price"] for i in all_data["Candy"] if i["price"] is not None]
    pop_sugars = [i for i in all_data["Pop"] if i["sugar"] is not None]
    top5_sugar = sorted(pop_sugars, key=lambda x: -x["sugar"])[:5]
    bot5_sugar = sorted(pop_sugars, key=lambda x: x["sugar"])[:5]
    top5_candy = sorted([i for i in all_data["Candy"] if i["price"] is not None],
                        key=lambda x: -x["price"])[:5]
    bot5_candy = sorted([i for i in all_data["Candy"] if i["price"] is not None],
                        key=lambda x: x["price"])[:5]

    cs = price_stats(candy_prices)

    cards_html = ""
    for name, items in all_data.items():
        cards_html += f'<div class="card"><div class="card-num">{len(items)}</div><div class="card-label">{esc(name)}</div></div>\n'

    def mini_table(headers, rows):
        h = "".join(f"<th>{esc(h)}</th>" for h in headers)
        body = ""
        for row in rows:
            body += "<tr>" + "".join(f"<td>{esc(str(c))}</td>" for c in row) + "</tr>"
        return f"<table class='mini'><thead><tr>{h}</tr></thead><tbody>{body}</tbody></table>"

    candy_expensive = mini_table(
        ["#", "Brand", "Variety", "$/4oz"],
        [(i + 1, r["brand"], r["variety"], f"${r['price']:.2f}") for i, r in enumerate(top5_candy)]
    )
    candy_cheapest = mini_table(
        ["#", "Brand", "Variety", "$/4oz"],
        [(i + 1, r["brand"], r["variety"], f"${r['price']:.2f}") for i, r in enumerate(bot5_candy)]
    )
    pop_highest = mini_table(
        ["#", "Brand", "Flavor", "Sugar(g)"],
        [(i + 1, r["brand"], r["flavor"], f"{r['sugar']:.0f}") for i, r in enumerate(top5_sugar)]
    )
    pop_lowest = mini_table(
        ["#", "Brand", "Flavor", "Sugar(g)"],
        [(i + 1, r["brand"], r["flavor"], f"{r['sugar']:.0f}") for i, r in enumerate(bot5_sugar)]
    )

    page_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Analytics — The CountUp</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: 'Courier New', Courier, monospace;
    font-weight: 700;
    background: #000;
    color: #fff;
    line-height: 1.4;
    padding: 20px;
  }}
  .container {{
    max-width: 1000px;
    margin: 0 auto;
    border: 3px solid #fff;
  }}
  header {{
    text-align: center;
    padding: 25px 20px 15px;
    background: #fff;
    color: #000;
    position: relative;
  }}
  header h1 {{
    font-family: Georgia, 'Times New Roman', serif;
    font-size: 2.4em;
    font-weight: 900;
    letter-spacing: 0.08em;
  }}
  header .subtitle {{
    font-family: 'Courier New', Courier, monospace;
    font-size: 0.85em;
    font-weight: 400;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    color: #555;
    margin-top: 4px;
  }}
  .back-link {{
    position: absolute;
    left: 16px;
    bottom: 14px;
    font-family: 'Courier New', Courier, monospace;
    font-size: 0.8em;
    font-weight: 700;
    color: #000;
    text-decoration: none;
    border-bottom: 1px solid #999;
  }}
  .back-link:hover {{ border-bottom-color: #000; }}
  main {{ padding: 20px; }}
  .cards {{
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
    justify-content: center;
    margin-bottom: 30px;
  }}
  .card {{
    border: 2px solid #fff;
    padding: 16px 20px;
    text-align: center;
    min-width: 130px;
  }}
  .card-num {{
    font-size: 2em;
    font-family: Georgia, serif;
  }}
  .card-label {{
    font-size: 0.75em;
    color: #888;
    margin-top: 4px;
    text-transform: uppercase;
    letter-spacing: 0.1em;
  }}
  .section-title {{
    font-family: Georgia, serif;
    font-size: 1.3em;
    margin: 30px 0 12px;
    border-bottom: 1px solid #333;
    padding-bottom: 6px;
  }}
  .stat-line {{
    font-size: 0.9em;
    color: #aaa;
    margin: 4px 0;
    font-weight: 400;
  }}
  .stat-line span {{ color: #fff; font-weight: 700; }}
  .chart-wrap {{
    text-align: center;
    margin: 20px 0;
  }}
  .chart-wrap img {{
    max-width: 100%;
    height: auto;
    border: 1px solid #333;
  }}
  table.mini {{
    width: 100%;
    border-collapse: collapse;
    margin: 10px 0 20px;
    font-size: 0.85em;
  }}
  table.mini th {{
    background: #222;
    padding: 8px 12px;
    text-align: left;
    font-weight: 700;
    border-bottom: 2px solid #555;
  }}
  table.mini td {{
    padding: 6px 12px;
    border-bottom: 1px solid #222;
    font-weight: 400;
  }}
  .two-col {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
  }}
  @media (max-width: 600px) {{
    .two-col {{ grid-template-columns: 1fr; }}
    .card {{ min-width: 100px; padding: 12px; }}
    .card-num {{ font-size: 1.5em; }}
  }}
  footer {{
    text-align: center;
    padding: 18px;
    border-top: 2px solid #fff;
    font-size: 0.7em;
    letter-spacing: 0.1em;
    color: #666;
  }}
</style>
</head>
<body>
<div class="container">
  <header>
    <a href="index.html" class="back-link">&larr; Menu</a>
    <h1>Analytics</h1>
    <div class="subtitle">{sum(len(v) for v in all_data.values())} total items ranked</div>
  </header>
  <main>
    <div class="cards">
{cards_html}
    </div>

    <h2 class="section-title">Candy Prices ($/4oz)</h2>
    <p class="stat-line">Avg: <span>${cs['avg']:.2f}</span> &middot; Median: <span>${cs['median']:.2f}</span> &middot; Range: <span>${cs['min']:.2f}</span> &ndash; <span>${cs['max']:.2f}</span></p>
    <div class="two-col">
      <div><h3 class="section-title">Most Expensive</h3>{candy_expensive}</div>
      <div><h3 class="section-title">Best Value</h3>{candy_cheapest}</div>
    </div>

    <h2 class="section-title">Pop Sugar (g per serving)</h2>
    <div class="two-col">
      <div><h3 class="section-title">Highest Sugar</h3>{pop_highest}</div>
      <div><h3 class="section-title">Lowest Sugar</h3>{pop_lowest}</div>
    </div>

    <h2 class="section-title">Items by Country / Origin</h2>
    <div class="chart-wrap">
      <img src="analytics_chart.png" alt="Items by country bar chart">
    </div>
  </main>
  <footer>The CountUp Rankings — Analytics</footer>
</div>
</body>
</html>"""

    path = os.path.join(ANALYTICS_DIR, "analytics.html")
    with open(path, "w", encoding="utf-8") as f:
        f.write(page_html)
    print(f"Written: {path}")
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Connecting to open workbook...")
    wb = xw.Book("TheCountUp Rankings.xlsx")

    print("Extracting data...")
    all_data = {
        "Games": extract_games(wb.sheets["Games"]),
        "Fast Food Hack": extract_dining(wb.sheets["Dine"]),
        "Pop": extract_pop(wb.sheets["Pop"]),
        "Candy": extract_candy(wb.sheets["Candy"]),
        "Chocolate": extract_chocolate(wb.sheets["Choc"]),
        "Sauces": extract_sauces(wb.sheets["Sauces"]),
    }
    for name, items in all_data.items():
        print(f"  {name}: {len(items)} items")

    print("\nGenerating Excel report...")
    write_excel(all_data)

    print("Generating chart...")
    make_chart(all_data)

    print("Generating HTML page...")
    build_html(all_data)

    print("\nDone.")


if __name__ == "__main__":
    main()
